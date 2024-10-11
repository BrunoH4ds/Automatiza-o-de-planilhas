import subprocess
import sys

try:
    from openpyxl import *
    from copy import copy
except ImportError:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', '--upgrade', 'pip'])
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'copy'])
finally:
    from openpyxl import *
    from copy import copy

def criar_aba(Bairro, arquivo):
    if Bairro not in arquivo.sheetnames:
        nova_aba = arquivo.create_sheet(Bairro)
        nova_aba = arquivo[Bairro]
        nova_aba['A1'].value = 'Data De Nascimento'
        nova_aba['B1'].value = 'Nome'
        nova_aba['C1'].value = 'Bairro'
        print(f'Aba "{Bairro}" criada com sucesso com cabeçalhos.')

def transferir_informacoes(aba_base, aba_destino, linha):
    linha_destino = aba_destino.max_row + 1
    for coluna in range(1, 4):
        celula_origem = aba_base.cell(row=linha, column=coluna)
        celula_destino = aba_destino.cell(row=linha_destino, column=coluna)
        celula_destino.value = celula_origem.value
        celula_destino._style = copy(celula_origem._style)

arquivo = load_workbook("Automatizacao excel\\Bairros.xlsx")
print(arquivo.sheetnames)

aba_base = arquivo['Base de Dados']

ultimalinha = aba_base.max_row
print(f"Última linha na aba base: {ultimalinha}")

for linha in range(2, ultimalinha + 1):
    Bairro = aba_base[f'C{linha}'].value
    if not Bairro:
        break

    criar_aba(Bairro, arquivo)
    aba_destino = arquivo[Bairro]
    transferir_informacoes(aba_base, aba_destino, linha)

arquivo.save('Automatizacao excel\\Bairros.xlsx')
