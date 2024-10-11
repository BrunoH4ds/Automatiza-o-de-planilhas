from openpyxl import load_workbook
from copy import copy

def criar_aba(bairro, arquivo_bairros, estilos_cabecalho):
    if bairro not in arquivo_bairros.sheetnames:
        nova_aba = arquivo_bairros.create_sheet(bairro)
        nova_aba["A1"].value = "Data de Nascimento"
        nova_aba["B1"].value = "Pessoa"
        nova_aba["C1"].value = "Bairro"
        # Aplicar estilos aos cabeçalhos
        nova_aba["A1"].font = estilos_cabecalho['font']
        nova_aba["A1"].alignment = estilos_cabecalho['alignment']
        nova_aba["A1"].border = estilos_cabecalho['border']
        nova_aba["B1"].font = estilos_cabecalho['font']
        nova_aba["B1"].alignment = estilos_cabecalho['alignment']
        nova_aba["B1"].border = estilos_cabecalho['border']
        nova_aba["C1"].font = estilos_cabecalho['font']
        nova_aba["C1"].alignment = estilos_cabecalho['alignment']
        nova_aba["C1"].border = estilos_cabecalho['border']

def transferir_informacoes_aba(aba_origem, aba_destino, linha_origem):
    linha_destino = aba_destino.max_row + 1
    for coluna in range(1, 4):
        celula_origem = aba_origem.cell(row=linha_origem, column=coluna)
        celula_destino = aba_destino.cell(row=linha_destino, column=coluna)
        celula_destino.value = celula_origem.value
        # Copiar estilo de maneira adequada
        celula_destino.font = copy(celula_origem.font)
        celula_destino.alignment = copy(celula_origem.alignment)
        celula_destino.border = copy(celula_origem.border)
        celula_destino.fill = copy(celula_origem.fill)

# Carregar o arquivo de bairros
arquivo_bairros = load_workbook("Automatizacao excel\\Bairros.xlsx")

print(arquivo_bairros.sheetnames)

aba_basedados = arquivo_bairros["Base de Dados"]

ultima_linha = aba_basedados.max_row
print(ultima_linha)

# Copiar o estilo dos cabeçalhos
estilos_cabecalho = {
    'font': copy(aba_basedados["A1"].font),
    'alignment': copy(aba_basedados["A1"].alignment),
    'border': copy(aba_basedados["A1"].border),
    'fill': copy(aba_basedados["A1"].fill)
}

for linha in range(2, ultima_linha + 1):
    bairro = aba_basedados.cell(row=linha, column=3).value
    if not bairro:
        break
    # Criar uma aba para o bairro
    criar_aba(bairro, arquivo_bairros, estilos_cabecalho)

    # Transferir as informações para a aba do bairro
    aba_destino = arquivo_bairros[bairro]
    transferir_informacoes_aba(aba_basedados, aba_destino, linha)

# Salvar as alterações no arquivo
arquivo_bairros.save("Automatizacao excel\\Bairros.xlsx")
